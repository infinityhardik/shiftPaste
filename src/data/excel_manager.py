"""Excel file handling for master clipboard items."""

from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime
import json
from PySide6.QtCore import QThread, Signal
import time

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    load_workbook = None
    Workbook = None

import pandas as pd


class ExcelWatcher(QThread):
    """Background thread that watches for Excel file changes."""

    file_changed = Signal(str)  # Emits filepath when changed

    def __init__(self, watch_directory: str):
        """Initialize Excel file watcher.

        Args:
            watch_directory: Directory to watch for Excel files
        """
        super().__init__()
        self.watch_directory = Path(watch_directory)
        self.running = False
        self.file_mtimes = {}  # Track modification times

    def run(self):
        """Start watching for file changes."""
        if not self.watch_directory.exists():
            self.watch_directory.mkdir(parents=True, exist_ok=True)

        self.running = True

        # Initialize file modification times
        for filepath in self.watch_directory.glob("*.xlsx"):
            try:
                self.file_mtimes[str(filepath)] = filepath.stat().st_mtime
            except OSError:
                pass

        # Poll for changes every 1 second
        while self.running:
            try:
                for filepath in self.watch_directory.glob("*.xlsx"):
                    filepath_str = str(filepath)
                    try:
                        current_mtime = filepath.stat().st_mtime
                        last_mtime = self.file_mtimes.get(filepath_str, 0)

                        # File was modified
                        if current_mtime > last_mtime:
                            self.file_mtimes[filepath_str] = current_mtime
                            # Small delay to ensure file is fully written
                            time.sleep(0.2)
                            self.file_changed.emit(filepath_str)
                    except OSError:
                        pass

            except Exception as e:
                print(f"Error in file watcher: {e}")

            time.sleep(1)  # Check every 1 second

    def stop(self):
        """Stop watching for file changes."""
        self.running = False


class ExcelManager:
    """Manages Excel file import/export and watching."""

    def __init__(self, master_directory: str = "data/Master"):
        """Initialize Excel manager.

        Args:
            master_directory: Path to directory with Excel master files
        """
        self.master_directory = Path(master_directory)
        self.master_directory.mkdir(parents=True, exist_ok=True)
        self.watcher = None
        self._setup_default_files()

    def _setup_default_files(self):
        """Create default Excel files if they don't exist."""
        default_categories = ["Pinned", "Work", "Personal"]

        for category in default_categories:
            filepath = self.master_directory / f"{category}.xlsx"

            if not filepath.exists():
                self._create_empty_excel(filepath, category)

    def _create_empty_excel(self, filepath: Path, category: str):
        """Create an empty Excel file with proper structure.

        Args:
            filepath: Path to create Excel file
            category: Category name
        """
        try:
            if Workbook is None:
                print("openpyxl not available, skipping Excel file creation")
                return

            wb = Workbook()
            ws = wb.active
            ws.title = category

            # Create headers
            headers = ["Content", "Timestamp", "Notes"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = cell.font.copy()
                cell.font = cell.font.copy()

            # Set column widths
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 30

            wb.save(filepath)
        except Exception as e:
            print(f"Error creating Excel file {filepath}: {e}")

    def import_from_excel(self, category: str) -> List[Dict[str, Any]]:
        """Import items from Excel file.

        Args:
            category: Category name

        Returns:
            List of items with 'content', 'timestamp', and 'notes'
        """
        filepath = self.master_directory / f"{category}.xlsx"

        if not filepath.exists():
            return []

        try:
            # Try using pandas first (more flexible)
            try:
                df = pd.read_excel(filepath, sheet_name=category)

                items = []
                for _, row in df.iterrows():
                    content = str(row.get('Content', '')).strip()
                    if not content or content.lower() == 'nan':
                        continue

                    timestamp_val = row.get('Timestamp', '')
                    notes = str(row.get('Notes', '')).strip()

                    # Parse timestamp
                    if isinstance(timestamp_val, (int, float)):
                        timestamp = int(timestamp_val)
                    elif isinstance(timestamp_val, str):
                        try:
                            dt = pd.to_datetime(timestamp_val)
                            timestamp = int(dt.timestamp())
                        except:
                            timestamp = int(datetime.now().timestamp())
                    else:
                        timestamp = int(datetime.now().timestamp())

                    items.append({
                        'content': content,
                        'timestamp': timestamp,
                        'notes': notes if notes and notes.lower() != 'nan' else ''
                    })

                return items

            except Exception as e:
                print(f"Error reading Excel with pandas: {e}")
                # Fallback to openpyxl
                if load_workbook is None:
                    return []

                wb = load_workbook(filepath)
                if category not in wb.sheetnames:
                    return []

                ws = wb[category]
                items = []

                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or not row[0]:  # Skip empty rows
                        continue

                    content = str(row[0]).strip()
                    if not content:
                        continue

                    timestamp = int(datetime.now().timestamp())
                    if len(row) > 1 and row[1]:
                        try:
                            if isinstance(row[1], str):
                                dt = datetime.fromisoformat(row[1])
                                timestamp = int(dt.timestamp())
                            elif isinstance(row[1], (int, float)):
                                timestamp = int(row[1])
                        except:
                            pass

                    notes = str(row[2]).strip() if len(row) > 2 and row[2] else ''

                    items.append({
                        'content': content,
                        'timestamp': timestamp,
                        'notes': notes
                    })

                return items

        except Exception as e:
            print(f"Error importing from Excel {filepath}: {e}")
            return []

    def export_to_excel(self, category: str, items: List[Dict[str, Any]]):
        """Export items to Excel file.

        Args:
            category: Category name
            items: List of items to export
        """
        filepath = self.master_directory / f"{category}.xlsx"

        try:
            if Workbook is None:
                print("openpyxl not available, skipping Excel export")
                return

            wb = Workbook()
            ws = wb.active
            ws.title = category

            # Create headers
            headers = ["Content", "Timestamp", "Notes"]
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num).value = header

            # Add items
            for row_num, item in enumerate(items, start=2):
                content = item.get('content', '')
                timestamp = item.get('timestamp', 0)
                notes = item.get('notes', '')

                # Convert timestamp to datetime string
                try:
                    dt = datetime.fromtimestamp(timestamp)
                    timestamp_str = dt.isoformat()
                except:
                    timestamp_str = datetime.now().isoformat()

                ws.cell(row=row_num, column=1).value = content
                ws.cell(row=row_num, column=2).value = timestamp_str
                ws.cell(row=row_num, column=3).value = notes

            wb.save(filepath)

        except Exception as e:
            print(f"Error exporting to Excel {filepath}: {e}")

    def get_all_categories(self) -> List[str]:
        """Get all available master categories.

        Returns:
            List of category names
        """
        categories = []

        for filepath in self.master_directory.glob("*.xlsx"):
            category = filepath.stem
            categories.append(category)

        return sorted(categories)

    def start_watching(self):
        """Start watching for external Excel file changes."""
        if not self.watcher:
            self.watcher = ExcelWatcher(str(self.master_directory))
            self.watcher.start()

    def stop_watching(self):
        """Stop watching for file changes."""
        if self.watcher:
            self.watcher.stop()
            self.watcher.wait()
            self.watcher = None

    def get_file_changed_signal(self):
        """Get the file_changed signal from watcher.

        Returns:
            Signal object or None
        """
        if self.watcher:
            return self.watcher.file_changed
        return None
